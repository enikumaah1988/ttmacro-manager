"""Excel 台帳の読み込み・行検証・行データ抽出。

このモジュールは ``openpyxl`` を直接使い、pandas には依存しない。
1 行は ``dict[str, Any]`` で表現し、空セルは ``None``。
"""

from __future__ import annotations

import ipaddress
import math
import re
from typing import Any, TypedDict

from openpyxl import load_workbook

from ttmacro.config import EXCEL_PATH, KEYS_DIR
from ttmacro.ttl_renderer import sanitize_name


class RowData(TypedDict):
    """``extract_row_data`` の戻り値型。

    Excel 1 行から TTL 生成に必要な情報をすべて含む。値はすべて
    文字列に正規化済み（空セルは空文字、空ポートは ``"22"`` がデフォルト）。
    """

    name: str
    host: str
    port: str
    user: str
    password: str
    keyfile_name: str
    post_cmd: str
    memo: str
    group1: str
    group2: str
    group3: str
    template: str


def is_blank(value: Any) -> bool:
    """セル値が空相当か判定する。

    ``None``、空文字（strip 後）、``float('nan')`` のいずれかなら True。

    Args:
        value: 判定対象。

    Returns:
        空相当なら True。
    """
    return (
        value is None
        or (isinstance(value, str) and value.strip() == "")
        or (isinstance(value, float) and math.isnan(value))
    )


def is_blank_row(row: dict[str, Any]) -> bool:
    """行の全セルが空相当なら True。"""
    return all(is_blank(v) for v in row.values())


def safe_str(val: Any) -> str:
    """セル値を文字列化、空相当なら空文字。"""
    if is_blank(val):
        return ""
    return str(val).strip()


def safe_get(row: dict[str, Any], key: str, default: str = "") -> str:
    """行 dict から値を取得し、空相当なら default を返す。

    Args:
        row: 行データ（列名 → セル値）。
        key: 取得するカラム名。
        default: 空相当時のフォールバック値。

    Returns:
        値の文字列表現（strip 済み）。
    """
    value = row.get(key)
    if is_blank(value):
        return default
    return str(value).strip()


def load_excel_data() -> tuple[list[str], list[dict[str, Any]]]:
    """Excel 台帳ファイルを読み込み、ヘッダと行データを返す。

    1 行目をヘッダとして扱い、それ以降を ``dict[列名, セル値]`` の
    リストとして返す（空白行も含む）。

    Returns:
        ``(headers, rows)`` のタプル。

    Raises:
        FileNotFoundError: ファイルが存在しない場合。
        PermissionError: ファイルが他のアプリで開かれている場合。
        ValueError: ファイルが空、またはヘッダ行が無い場合。
        RuntimeError: その他の読み込みエラー。
    """
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {EXCEL_PATH}")

    try:
        # data_only=True で数式の評価結果を取得（生の数式ではなく）
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            raise ValueError("Excelファイルにアクティブなシートがありません")

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            wb.close()
            raise ValueError("Excelファイルが空です") from None

        # 末尾の None（Excel が空列をパディングする）を切り捨てる
        headers = list(header_row)
        while headers and headers[-1] is None:
            headers.pop()
        if not headers:
            wb.close()
            raise ValueError("Excelファイルにヘッダ行がありません")

        rows: list[dict[str, Any]] = []
        for row_values in rows_iter:
            row_dict: dict[str, Any] = {}
            for h, v in zip(headers, row_values, strict=False):
                if h is None:
                    continue  # 中間 None ヘッダは無視（防御）
                row_dict[h] = v
            rows.append(row_dict)

        wb.close()

        if not rows:
            raise ValueError("Excelファイルにデータ行がありません")

        # str 型のヘッダのみ返却（型ヒントを満たすため）
        return [str(h) for h in headers if h is not None], rows
    except PermissionError as e:
        raise PermissionError(f"Excelファイルが他で開かれています: {EXCEL_PATH}") from e
    except (FileNotFoundError, ValueError):
        raise
    except Exception as e:
        raise RuntimeError(f"Excelファイル読み込みエラー: {e}") from e


def validate_row_data(row: dict[str, Any], row_num: int) -> tuple[bool, list[str]]:
    """行データの妥当性を検証する。

    必須項目（name/host/user）、ホスト名/IP の形式、ポート番号の範囲、
    keyfile の存在をチェックする。

    Args:
        row: 検証対象の行（dict）。
        row_num: ログ用の行番号（現状未使用、将来用に保持）。

    Returns:
        ``(is_valid, errors)`` のタプル。
    """
    errors: list[str] = []

    # 必須フィールドチェック
    required_fields = ["name", "host", "user"]
    for field in required_fields:
        if is_blank(row.get(field)):
            errors.append(f"必須項目 '{field}' が空です")

    # IPアドレス/ホスト名チェック
    host = safe_str(row.get("host"))
    if host:
        try:
            ipaddress.ip_address(host)
        except ValueError:
            # IP として無効ならホスト名扱い（簡易チェック）
            if not re.match(r"^[a-zA-Z0-9.-]+$", host):
                errors.append(f"ホスト名 '{host}' の形式が不正です")

    # ポート番号チェック
    port = row.get("port")
    if not is_blank(port):
        try:
            port_num = int(port)  # type: ignore[arg-type]
            if not (1 <= port_num <= 65535):
                errors.append(f"ポート番号 {port_num} は範囲外です (1-65535)")
        except (ValueError, TypeError):
            errors.append(f"ポート番号 '{port}' が数値ではありません")

    # キーファイル存在チェック
    keyfile = safe_get(row, "keyfile")
    if keyfile:
        keyfile_path = KEYS_DIR / keyfile
        if not keyfile_path.exists():
            errors.append(f"キーファイル '{keyfile}' が見つかりません: {keyfile_path}")

    return len(errors) == 0, errors


def extract_row_data(row: dict[str, Any]) -> RowData:
    """行データから TTL 生成に必要な情報を抽出する。

    Args:
        row: 抽出元の行（dict）。

    Returns:
        ``RowData`` TypedDict（全フィールド正規化済み文字列）。
    """
    # メモ内の改行・タブは半角空白に置換（TTL コメントが壊れないように）
    memo = (
        safe_get(row, "memo").replace("\r", " ").replace("\n", " ").replace("\t", " ")
    )

    port_value = row.get("port")
    port_str = str(int(port_value)) if not is_blank(port_value) else "22"  # type: ignore[arg-type]

    return {
        "name": sanitize_name(safe_str(row.get("name"))),
        "host": safe_str(row.get("host")),
        "port": port_str,
        "user": safe_str(row.get("user")),
        "password": safe_get(row, "password"),
        "keyfile_name": safe_get(row, "keyfile"),
        "post_cmd": safe_get(row, "post_cmd"),
        "memo": memo,
        "group1": safe_get(row, "group1"),
        "group2": safe_get(row, "group2"),
        "group3": safe_get(row, "group3"),
        "template": safe_get(row, "template"),  # 空ならデフォルトテンプレ
    }
